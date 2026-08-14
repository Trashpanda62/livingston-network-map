"""Build the website-prospect call sheet (xlsx) from data/prospects.json.

One row per prospect — the 39 swept-up no/weak-website businesses plus the
flagged already-listed nodes — sorted by Google review count (demand proxy),
with phone, address, and the reason they need a site. Pure data table, no
formulas. Output goes to Downloads for immediate use.

Run: py -X utf8 scripts/build_call_sheet.py
"""
import json
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT = Path(r"C:\Users\Steve\Downloads\livingston-website-prospects.xlsx")

REASONS = {
    "no-website": "No website at all",
    "facebook-only": "Facebook page only",
    "rented-subdomain": "Rented site-builder page",
    "aggregator-listing": "Only on other people's sites",
    "dead-site": "Has a domain, effectively no visitors",
}


def host_of(url):
    h = (urlparse(url or "").hostname or "").lower()
    return h[4:] if h.startswith("www.") else h


def main() -> int:
    pros = json.loads((ROOT / "data" / "prospects.json").read_text(encoding="utf-8"))
    net = json.loads((ROOT / "data" / "network.json").read_text(encoding="utf-8"))
    nodes = {n["id"]: n for n in net["nodes"]}

    rows = []
    for r in pros.get("new", []):
        rows.append({
            "name": r["name"], "town": r.get("city") or "",
            "category": r.get("category") or "",
            "phone": r.get("phone") or "",
            "reviews": r.get("reviews"), "rating": r.get("rating"),
            "why": REASONS.get(r["reason"], r["reason"]),
            "presence": {"facebook-only": "Facebook page",
                         "rented-subdomain": "site-builder subdomain"}.get(r["reason"], "none found"),
            "address": r.get("address") or "",
            "on_map": "new find (not in our directories yet)",
        })
    for pid, f in pros.get("flagged", {}).items():
        n = nodes.get(pid)
        if not n:
            continue
        rows.append({
            "name": n["name"], "town": n.get("city") or "",
            "category": n.get("category") or "",
            "phone": f.get("phone") or "",
            "reviews": f.get("reviews"), "rating": None,
            "why": REASONS.get(f["reason"], f["reason"]),
            "presence": host_of(n.get("url")),
            "address": f.get("address") or "",
            "on_map": "already listed on our directories",
        })
    rows.sort(key=lambda r: -(r["reviews"] or 0))

    wb = Workbook()
    ws = wb.active
    ws.title = "Call list"
    headers = ["#", "Business", "Town", "Category", "Phone", "Google reviews",
               "Rating", "Why they need a site", "Current web presence",
               "Address", "Relationship", "Call notes"]
    ws.append(headers)
    head_font = Font(name="Arial", bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1F3864")
    for c in ws[1]:
        c.font = head_font
        c.fill = head_fill
        c.alignment = Alignment(vertical="center")
    for i, r in enumerate(rows, 1):
        ws.append([i, r["name"], r["town"], r["category"], r["phone"],
                   r["reviews"], r["rating"], r["why"], r["presence"],
                   r["address"], r["on_map"], ""])
    body = Font(name="Arial", size=10)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = body
    widths = [4, 34, 12, 16, 16, 13, 7, 28, 26, 40, 30, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:L{len(rows) + 1}"
    ws.freeze_panes = "A2"

    OUT.parent.mkdir(exist_ok=True)
    wb.save(OUT)
    with_phone = sum(1 for r in rows if r["phone"])
    print(f"wrote {OUT} — {len(rows)} rows, {with_phone} with phone")
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
